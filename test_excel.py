import openpyxl
import datetime
import os
from openpyxl.utils import get_column_letter, column_index_from_string

user = "邓杰"
condition = "BGM3"
wafer_id = "T1AG87.21"
marking = "AD0018AU2-ES"
dut_id = 10
pm_pos_start_list = {'1.56': 'L20', '1.6': 'L48', '1.64': 'L83', '1.68': 'L113', '1.72': 'L144'}
nonce_pos_start_list = {'1.56': 'H26', '1.6': 'H59', '1.64': 'H89', '1.68': 'H119', '1.72': 'H150'}
file_path = "data/"
isExists=os.path.exists(file_path)
if not isExists:
    # 如果不存在则创建目录
    # 创建目录操作函数
    os.makedirs(file_path)
    print(file_path + ' 创建成功')
else:
    # 如果目录存在则不创建，并提示目录已存在
    print(file_path + ' 目录已存在')

base_options = "--bmsc-options 115200:0 -o stratum.f2pool.com:3333 -u xuelei5151.1 -T --set-dc-com COM6:9600 --pattern "
special_options = "--bmsc-rdreg 0000 --bmsc-vil 1 --bmsc-ctype 1390 --bmsc-blken 52:1 --bmsc-domain 0 --bmsc-domain-cross "
log_options = "--logfile 1393.log --logfile-openflag w "
extra_options = "--bmsc-setpll 1390:{freq} --cp2103-reset-chip --bmsc-scan-freq-step {scan_freq_step} --bmsc-scan-freq-nonce-num {scan_freq_nonce_num} ".format(
    freq=200,
    scan_freq_nonce_num=180,
    scan_freq_step=50
)

for volt in pm_pos_start_list.keys():
    volt_option = "--bmsc-voltage {volt_value} ".format(volt_value=float(volt))
    run_cgminer_str = "cgminer-1393.exe " + base_options + special_options + extra_options + volt_option + log_options
    nonce_state_name = file_path + "nonce_state_" + volt
    nonce_state_cur_name = file_path + "nonce_state_cur_" + volt
    os.system(run_cgminer_str)
    os.system(
        "cat 1393.log | grep ':golden nonce' | awk '{ printf $5\"\\t\"$9 \"\\n\"}' | awk -F, '{ printf $1\"\\n\" }' | awk '{nonce[$1]=nonce[$1](\"\\t\"$2);sum[$1]+=$2}END{for(i in nonce)print i,\"\\t\"sum[i],nonce[i]}' | sort > " + nonce_state_name)
    os.system(
        "cat 1393.log | grep 'Before open core, Get DC current is'| awk '{printf \"%.4f\\n\",$10}' > data/leakage_cur && dos2unix data/leakage_cur")
    os.system(
        "cat 1393.log | grep 'After test pattern, Get DC current is'| awk '{printf \"%.4f\\n\",$10}' > data/state_cur && dos2unix data/state_cur")
    os.system("paste -d \"\\t\" data/leakage_cur data/state_cur " + nonce_state_name + "> " + nonce_state_cur_name)

wb = openpyxl.load_workbook('test.xlsx')
sheet = wb.copy_worksheet(wb.worksheets[0])

for volt in pm_pos_start_list.keys():
    fopen = open("pm_data_" + volt, 'r')
    pm_lines = fopen.readlines()
    fopen.close()
    fopen = open("nonce_state_cur_" + volt, 'r')
    nonce_state_lines = fopen.readlines()
    fopen.close()

    i = 0
    for line in pm_lines:
        li = line.split('\t')
        val = int(li[0])
        sheet.cell(int(pm_pos_start_list[volt][1:]), column_index_from_string(pm_pos_start_list[volt][:1]) + i, val)
        i = i + 1

    i = 0
    for line in nonce_state_lines:
        li = line.split('\t')
        j = 0
        for l in li:
            val = float(l)
            sheet.cell(int(nonce_pos_start_list[volt][1:]) + i,
                       column_index_from_string(nonce_pos_start_list[volt][:1]) + j, val)
            j = j + 1
        i = i + 1

sheet['B22'] = 0.4
sheet['C22'] = 0.8
sheet['D22'] = 1.2

wafer_info = wafer_id.split('.')
date_str = str(datetime.datetime.today())

sheet['B3'] = user
sheet['E3'] = date_str[:16]
sheet['A5'] = "1 芯片信息:(" + condition + ")"
sheet['K7'] = wafer_info[0]
sheet['L7'] = int(wafer_info[1])
sheet['B14'] = "(" + str(dut_id) + "号板)"
sheet.title = wafer_id + "#" + str(dut_id)

# wb.save("BM1393测试_"+condition+"_SLT_"+date_str[5:10]+".xlsx")
wb.save('test.xlsx')


